import streamlit as st
import pandas as pd
from rapidfuzz import fuzz
import unicodedata
import re
from io import BytesIO
from datetime import time
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

st.set_page_config(layout="wide")
st.title("📋 Backpage ↔ DTS Cross Check")

# -------------------------------------------------
# Uploads
# -------------------------------------------------
backpage_file = st.file_uploader("Upload Backpage Excel", type="xlsx")
sf_file = st.file_uploader("Upload SF.xlsx", type="xlsx")
dts_file = st.file_uploader("Upload DTS.xlsx", type="xlsx")

# -------------------------------------------------
# Helpers
# -------------------------------------------------
def normalize(text):
    if pd.isna(text):
        return ""
    text = str(text).strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(c for c in text if not unicodedata.combining(c))
    text = re.sub(r"#\s*(\d+)", r"\1", text)
    text = re.sub(r"\s+", " ", text)
    return text

def compare_times(call, start):
    if isinstance(call, time) and isinstance(start, time):
        if call == start:
            return "Match"
        elif call < start:
            return "Later Start"
        else:
            return "Earlier Start"
    return ""

# -------------------------------------------------
# Processing
# -------------------------------------------------
if backpage_file and sf_file and dts_file:
    if st.button("▶️ Process Files"):
        with st.spinner("Processing..."):

            # ---------- Backpage ----------
            df_raw = pd.read_excel(backpage_file, header=None)

            def extract_block(df, col_start, col_end):
                block = df.iloc[0:200, col_start:col_end].copy()
                block.columns = ["C", "Title", "Name", "Call"]
                return block

            df_backpage = pd.concat([
                extract_block(df_raw, 1, 5),
                extract_block(df_raw, 5, 9),
                extract_block(df_raw, 9, 13)
            ], ignore_index=True)

            df_backpage = df_backpage.dropna(subset=["Title", "Name"])
            df_backpage = df_backpage[
                (df_backpage["Title"].astype(str).str.strip() != "") &
                (df_backpage["Name"].astype(str).str.strip() != "")
            ]

            # ---------- SF matching ----------
            df_sf = pd.read_excel(sf_file)
            sf_pairs = list(zip(
                df_sf["Crew_list_name"],
                df_sf["Job_title"],
                df_sf["Sf_number"]
            ))

            def find_best_match(name, title):
                best_score, best_sf = 0, None
                for sf_name, sf_title, sf_number in sf_pairs:
                    score = (
                        0.7 * fuzz.token_set_ratio(normalize(name), normalize(sf_name)) +
                        0.3 * fuzz.token_sort_ratio(normalize(title), normalize(sf_title))
                    )
                    if score > best_score:
                        best_score = score
                        best_sf = sf_number
                return best_sf if best_score >= 85 else None

            df_backpage["SF ID"] = df_backpage.apply(
                lambda r: find_best_match(r["Name"], r["Title"]),
                axis=1
            )

            # ---------- DTS ----------
            df_dts = pd.read_excel(dts_file)[["Name", "Title", "Start", "SF"]].copy()
            df_dts.rename(columns={"SF": "SF ID"}, inplace=True)

            dts_by_sf = df_dts.set_index("SF ID")
            bp_by_sf = df_backpage.set_index("SF ID")

            # ---------- Backpage columns ----------
            def start_on_dts(row):
                sf_id = row["SF ID"]
                if pd.isna(sf_id) or str(sf_id).strip() == "":
                    return "", ""

                if sf_id not in dts_by_sf.index:
                    return "", "Missing"

                start_val = dts_by_sf.loc[sf_id, "Start"]
                if isinstance(start_val, pd.Series):
                    start_val = start_val.iloc[0]

                if pd.isna(start_val):
                    return "", "Worked"

                return start_val, compare_times(row["Call"], start_val)

            df_backpage[["DTS Start", "Start on DTS"]] = df_backpage.apply(
                lambda r: pd.Series(start_on_dts(r)), axis=1
            )

            # ---------- DTS columns ----------
            def call_on_backpage(row):
                sf_id = row["SF ID"]
                if sf_id not in bp_by_sf.index:
                    return "", "-"

                call_val = bp_by_sf.loc[sf_id, "Call"]
                if isinstance(call_val, pd.Series):
                    call_val = call_val.iloc[0]

                return call_val, compare_times(call_val, row["Start"])

            df_dts[["Backpage Call", "Call on Backpage"]] = df_dts.apply(
                lambda r: pd.Series(call_on_backpage(r)), axis=1
            )

            # ---------- Excel output ----------
            output = BytesIO()
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                df_backpage.to_excel(writer, sheet_name="Backpage", index=False)
                df_dts.to_excel(writer, sheet_name="DTS", index=False)

            wb = load_workbook(output)

            for sheet in ["Backpage", "DTS"]:
                ws = wb[sheet]
                for col in ws.columns:
                    width = max(len(str(c.value)) if c.value else 0 for c in col)
                    ws.column_dimensions[get_column_letter(col[0].column)].width = min(width + 2, 50)

            wb.save(output)

            st.success("✅ Done")

            st.download_button(
                "📥 Download Excel",
                data=output.getvalue(),
                file_name="Backpage_DTS_Check.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
